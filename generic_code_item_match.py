"""
Generic Description Matching Engine
=====================================
Matches item descriptions between two customer/institution datasets.
Handles messy medical supply, device, and pharmacy descriptions with
comprehensive auto-cleaning and multi-strategy matching.

Works for: Medline, Boston Scientific, pharmacy, devices, supplies, etc.

Requirements:
    pip install pandas openpyxl rapidfuzz scikit-learn numpy

Usage:
    1. Update CONFIG section below (file paths, column names)
    2. Run:  python generic_description_matching.py
"""

import pandas as pd
import re
import numpy as np
from rapidfuzz import fuzz
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')


# ============================================================
# CONFIG — update these for your data
# ============================================================
FILE_A = "cchs_desc.xlsx"                    # Customer A file
FILE_B = "Item_description_umass.xlsx"       # Customer B file
OUTPUT_FILE = "description_matching_results.xlsx"

COL_A = "Line_desc_cchs"                    # Column name in file A
COL_B = "Item_description_umass"             # Column name in file B

LABEL_A = "Customer_A"                       # Display label (no real names)
LABEL_B = "Customer_B"                       # Display label

# File format (for .xlsx use 'excel', for .csv/.txt use 'csv')
FILE_A_FORMAT = "excel"                      # 'excel' or 'csv'
FILE_B_FORMAT = "excel"                      # 'excel' or 'csv'
CSV_SEP = "|"                                # Separator if csv
CSV_ENCODING = "utf-8"

# Matching parameters
TFIDF_THRESHOLD = 0.10           # Min TF-IDF score to consider
COMBINED_MIN_SCORE = 30          # Min combined score to keep
TOP_K_CANDIDATES = 5             # Candidates per item before 1:1
WORD_TFIDF_WEIGHT = 0.6          # Weight for word n-gram TF-IDF
CHAR_TFIDF_WEIGHT = 0.4          # Weight for char n-gram TF-IDF
TFIDF_WEIGHT = 0.5               # Weight for TF-IDF in final score
FUZZY_WEIGHT = 0.5               # Weight for fuzzy in final score
BATCH_SIZE = 300                 # Batch size for similarity computation

# Confidence thresholds
HIGH_THRESHOLD = 70
MEDIUM_THRESHOLD = 50
LOW_THRESHOLD = 35

# Abbreviations dictionary file (place in same folder as this script)
# If not found, built-in defaults are used
ABBREVIATIONS_FILE = "medical_abbreviations.txt"
# ============================================================


# ============================================================
# GENERIC DATA CLEANING ENGINE
# ============================================================

class DescriptionCleaner:
    """Auto-detects and removes noise patterns from healthcare descriptions."""

    # All known noise patterns across medical supply datasets
    CLEANING_RULES = [
        # --- Prefix patterns ---
        (r'ITM-\d+\s*-\s*', ''),                    # ITM-401592 - ...
        (r'\[\s*Catalog\s*:.*?\]', ''),              # [ Catalog: NPKP700RF ]
        (r'\[\s*Item\s*:.*?\]', ''),                 # [ Item: ITM-439823 ]
        (r'^\d{5,}\|', ''),                          # 0033673|text
        (r'^DIV\d+/\s*Various\s+Components\s*:\s*', ''),  # DIV22/Various Components:
        (r'^MPB\s+', ''),                            # MPB STELARA 130MG...

        # --- Suffix patterns ---
        (r'@+\s*$', ''),                             # Trailing @ (pharmacy pack codes)

        # --- Embedded noise ---
        (r'\bDYNJ\w+', ''),                          # Medline catalog codes (DYNJ902424F)
        (r'\b[A-Z]{2,4}\d{6,}\b', ''),              # Generic alphanumeric catalog codes
    ]

    # Medical abbreviation expansions for better matching
    # Loaded from external file if available, otherwise uses built-in defaults
    ABBREVIATIONS = {}  # Populated by _load_abbreviations()

    # Built-in fallback (used if external file not found)
    _DEFAULT_ABBREVIATIONS = {
        r'\bCATH\b': 'CATHETER',
        r'\bDIL\b': 'DILATATION',
        r'\bPTCA\b': 'ANGIOPLASTY',
        r'\bOTW\b': 'OVER THE WIRE',
        r'\bGW\b': 'GUIDEWIRE',
        r'\bSTRL\b': 'STERILE',
        r'\bSTER\b': 'STERILE',
        r'\bDISP\b': 'DISPOSABLE',
        r'\bSURG\b': 'SURGICAL',
        r'\bSYN\b': 'SYNTHETIC',
        r'\bABD\b': 'ABDOMINAL',
        r'\bPWDR\b': 'POWDER',
        r'\bSOL\b': 'SOLUTION',
        r'\bINJ\b': 'INJECTION',
        r'\bTB\b': 'TABLET',
        r'\bCP\b': 'CAPSULE',
        r'\bSYR\b': 'SYRINGE',
        r'\bNDL\b': 'NEEDLE',
        r'\bBX\b': 'BOX',
        r'\bCS\b': 'CASE',
        r'\bPK\b': 'PACK',
        r'\bLG\b': 'LARGE',
        r'\bSM\b': 'SMALL',
        r'\bMED\b': 'MEDIUM',
        r'\bXL\b': 'EXTRA LARGE',
        r'\bRT\b': 'RIGHT',
        r'\bLT\b': 'LEFT',
        r'\bLF\b': 'LATEX FREE',
    }

    def __init__(self, expand_abbreviations=True, abbreviations_file='medical_abbreviations.txt'):
        self.expand_abbreviations = expand_abbreviations
        self._detected_patterns = {}
        self._load_abbreviations(abbreviations_file)

    def _load_abbreviations(self, filepath):
        """Load abbreviations from external text file. Falls back to built-in defaults."""
        import os
        if os.path.exists(filepath):
            loaded = {}
            with open(filepath, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    # Skip comments and blank lines
                    if not line or line.startswith('#'):
                        continue
                    if '=' in line:
                        parts = line.split('=', 1)
                        abbr = parts[0].strip()
                        expansion = parts[1].strip()
                        if abbr and expansion:
                            # Wrap with word boundary regex
                            loaded[r'\b' + abbr + r'\b'] = expansion
            self.ABBREVIATIONS = loaded
            print(f"  Loaded {len(loaded)} abbreviations from {filepath}")
        else:
            self.ABBREVIATIONS = self._DEFAULT_ABBREVIATIONS.copy()
            print(f"  Abbreviation file not found ({filepath}), using {len(self.ABBREVIATIONS)} built-in defaults")

    def detect_patterns(self, descriptions):
        """Auto-detect which noise patterns exist in this dataset."""
        sample = [str(d) for d in descriptions[:2000]]
        self._detected_patterns = {}

        checks = {
            'has_itm_prefix': (r'ITM-\d+', 50),
            'has_catalog_bracket': (r'\[\s*Catalog', 20),
            'has_item_bracket': (r'\[\s*Item', 10),
            'has_numeric_pipe': (r'^\d{5,}\|', 20),
            'has_div_prefix': (r'^DIV\d+/', 2),
            'has_trailing_at': (r'@\s*$', 50),
            'has_mpb_prefix': (r'^MPB\s', 20),
            'has_duplicated_text': (None, 10),    # Special check
            'has_product_colon_detail': (r'^[A-Za-z\s]+:', 50),
        }

        for name, (pattern, threshold) in checks.items():
            if name == 'has_duplicated_text':
                count = sum(1 for d in sample
                            if len(d) > 30 and d[:len(d)//2].strip() == d[len(d)//2:].strip())
            else:
                count = sum(1 for d in sample if re.search(pattern, d, re.IGNORECASE))
            self._detected_patterns[name] = count >= threshold

        return self._detected_patterns

    def _remove_duplicated_text(self, text):
        """Fix descriptions where text is repeated: 'ABC DEFABC DEF' -> 'ABC DEF'."""
        if len(text) < 20:
            return text
        half = len(text) // 2
        # Try exact half split
        if text[:half].strip() == text[half:].strip():
            return text[:half].strip()
        # Try with 1-2 char tolerance
        for offset in range(-3, 4):
            mid = half + offset
            if mid > 10 and mid < len(text) - 10:
                left = text[:mid].strip()
                right = text[mid:].strip()
                if left == right:
                    return left
                # Check if right starts with left (truncated duplicate)
                if len(left) > 20 and right.startswith(left[:20]):
                    return left
        return text

    def clean(self, text):
        """Apply all cleaning rules to a description."""
        text = str(text).strip()
        if not text or text.lower() == 'nan':
            return ''

        # Uppercase for consistency
        text = text.upper()

        # Apply regex cleaning rules
        for pattern, replacement in self.CLEANING_RULES:
            text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)

        # Remove duplicated text
        text = self._remove_duplicated_text(text)

        # Normalize dimensions: "4.0X20" -> "4.0 X 20"
        text = re.sub(r'(\d)X(\d)', r'\1 X \2', text)
        text = re.sub(r'(\d)MM\b', r'\1 MM', text)
        text = re.sub(r'(\d)CM\b', r'\1 CM', text)
        text = re.sub(r'(\d)IN\b', r'\1 IN', text)
        text = re.sub(r'(\d)FR\b', r'\1 FR', text)

        # Expand abbreviations if enabled
        if self.expand_abbreviations:
            for pattern, expansion in self.ABBREVIATIONS.items():
                text = re.sub(pattern, expansion, text)

        # Final whitespace cleanup
        text = re.sub(r'[^\w\s/\.\-\,]', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip()

        return text


# ============================================================
# MATCHING ENGINE
# ============================================================

class DescriptionMatcher:
    """Multi-strategy description matching with 1:1 greedy assignment."""

    def __init__(self, config=None):
        self.config = config or {}

    def build_tfidf(self, descs_a, descs_b):
        """Build dual TF-IDF models (word + char n-grams)."""
        all_descs = descs_a + descs_b

        # Word n-grams: better for product name matching
        vec_word = TfidfVectorizer(
            analyzer='word', ngram_range=(1, 3),
            max_features=80000, sublinear_tf=True, min_df=1
        )
        tfidf_word = vec_word.fit_transform(all_descs)

        # Char n-grams: catches abbreviations and partial matches
        vec_char = TfidfVectorizer(
            analyzer='char_wb', ngram_range=(3, 5),
            max_features=80000, sublinear_tf=True
        )
        tfidf_char = vec_char.fit_transform(all_descs)

        n_a = len(descs_a)
        return (tfidf_word[:n_a], tfidf_word[n_a:],
                tfidf_char[:n_a], tfidf_char[n_a:])

    def compute_matches(self, descs_a_clean, descs_b_clean):
        """Find best matches using combined TF-IDF + fuzzy scoring."""

        print("  Building TF-IDF models (word + char)...")
        a_word, b_word, a_char, b_char = self.build_tfidf(descs_a_clean, descs_b_clean)

        print("  Computing similarity scores...")
        candidates = []
        total = len(descs_a_clean)

        for i in range(0, total, BATCH_SIZE):
            end = min(i + BATCH_SIZE, total)

            sim_word = cosine_similarity(a_word[i:end], b_word)
            sim_char = cosine_similarity(a_char[i:end], b_char)
            sim = WORD_TFIDF_WEIGHT * sim_word + CHAR_TFIDF_WEIGHT * sim_char

            for j in range(sim.shape[0]):
                a_idx = i + j
                top_indices = np.argsort(sim[j])[-TOP_K_CANDIDATES:][::-1]

                for rank, b_idx in enumerate(top_indices):
                    score = sim[j][b_idx]
                    if score > TFIDF_THRESHOLD:
                        # Multi-fuzzy: token sort + partial ratio
                        fz_token = fuzz.token_sort_ratio(
                            descs_a_clean[a_idx], descs_b_clean[b_idx]
                        )
                        fz_partial = fuzz.partial_ratio(
                            descs_a_clean[a_idx], descs_b_clean[b_idx]
                        )
                        fz_best = max(fz_token, fz_partial)

                        combined = TFIDF_WEIGHT * score * 100 + FUZZY_WEIGHT * fz_best

                        if combined >= COMBINED_MIN_SCORE:
                            candidates.append({
                                'a_idx': a_idx,
                                'b_idx': b_idx,
                                'tfidf': round(score * 100, 1),
                                'fuzzy': round(fz_best, 1),
                                'combined': round(combined, 1),
                                'rank': rank + 1
                            })

            processed = min(i + BATCH_SIZE, total)
            if processed % 1000 == 0 or processed == total:
                print(f"    {processed}/{total}...")

        print(f"  Total candidate pairs: {len(candidates)}")
        return candidates

    def greedy_1to1(self, candidates):
        """Assign 1:1 matches greedily by highest combined score."""
        df = pd.DataFrame(candidates)
        if df.empty:
            return []

        df.sort_values('combined', ascending=False, inplace=True)

        used_a, used_b = set(), set()
        matches = []

        for _, row in df.iterrows():
            ai, bi = int(row['a_idx']), int(row['b_idx'])
            if ai not in used_a and bi not in used_b:
                matches.append(row.to_dict())
                used_a.add(ai)
                used_b.add(bi)

        return matches


# ============================================================
# EXCEL WRITER
# ============================================================

def write_results_excel(descs_a, descs_b, descs_a_clean, descs_b_clean,
                        matches, all_candidates, output_path):
    """Write formatted Excel with Summary, Best Matches, Top-3, Unmatched."""

    def confidence(score):
        if score >= HIGH_THRESHOLD: return 'High'
        if score >= MEDIUM_THRESHOLD: return 'Medium'
        if score >= LOW_THRESHOLD: return 'Low'
        return 'Very Low'

    # Build results DataFrame
    rows = []
    for m in matches:
        ai, bi = int(m['a_idx']), int(m['b_idx'])
        rows.append({
            f'{LABEL_A}_Description': descs_a[ai],
            f'{LABEL_A}_Cleaned': descs_a_clean[ai],
            f'{LABEL_B}_Description': descs_b[bi],
            f'{LABEL_B}_Cleaned': descs_b_clean[bi],
            'TF-IDF_Score': m['tfidf'],
            'Fuzzy_Score': m['fuzzy'],
            'Combined_Score': m['combined'],
            'Confidence': confidence(m['combined']),
        })

    df_best = pd.DataFrame(rows).sort_values('Combined_Score', ascending=False)

    # Top-3 candidates
    df_cand = pd.DataFrame(all_candidates)
    if not df_cand.empty:
        df_top3 = (df_cand.sort_values(['a_idx', 'combined'], ascending=[True, False])
                   .groupby('a_idx').head(3).reset_index(drop=True))
        top3_rows = []
        for _, r in df_top3.iterrows():
            ai, bi = int(r['a_idx']), int(r['b_idx'])
            top3_rows.append({
                f'{LABEL_A}_Description': descs_a[ai],
                f'{LABEL_B}_Description': descs_b[bi],
                'TF-IDF_Score': r['tfidf'],
                'Fuzzy_Score': r['fuzzy'],
                'Combined_Score': r['combined'],
                'Match_Rank': r['rank'],
            })
        df_top3_out = pd.DataFrame(top3_rows)
    else:
        df_top3_out = pd.DataFrame()

    # --- Excel ---
    wb = Workbook()
    hfill = PatternFill('solid', fgColor='1F4E79')
    hfont = Font(bold=True, color='FFFFFF', size=10)
    thin = Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'), bottom=Side('thin'))
    conf_colors = {
        'High': PatternFill('solid', fgColor='C6EFCE'),
        'Medium': PatternFill('solid', fgColor='FFEB9C'),
        'Low': PatternFill('solid', fgColor='FFC7CE'),
        'Very Low': PatternFill('solid', fgColor='D9D9D9'),
    }

    # --- Summary ---
    ws0 = wb.active
    ws0.title = 'Summary'
    ws0['A1'] = 'Description Matching Report'
    ws0['A1'].font = Font(bold=True, size=14)
    summary = [
        ('', ''),
        (f'{LABEL_A} Descriptions', len(descs_a)),
        (f'{LABEL_B} Descriptions', len(descs_b)),
        ('Matched Pairs (1:1 unique)', len(df_best)),
        ('', ''),
        ('CONFIDENCE BREAKDOWN', ''),
    ]
    for conf in ['High', 'Medium', 'Low', 'Very Low']:
        summary.append((conf, int((df_best['Confidence'] == conf).sum()) if len(df_best) > 0 else 0))

    for r, (label, val) in enumerate(summary, 3):
        ws0[f'A{r}'] = label
        ws0[f'B{r}'] = val
        if label and label.isupper():
            ws0[f'A{r}'].font = Font(bold=True)
    ws0.column_dimensions['A'].width = 32
    ws0.column_dimensions['B'].width = 15

    # --- Best Matches ---
    ws1 = wb.create_sheet('Best Matches')
    best_headers = [
        (f'{LABEL_A}_Description', 55),
        (f'{LABEL_B}_Description', 55),
        ('TF-IDF_Score', 12),
        ('Fuzzy_Score', 12),
        ('Combined_Score', 14),
        ('Confidence', 12),
    ]

    for c, (h, w) in enumerate(best_headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.border = hfont, hfill, thin
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws1.column_dimensions[get_column_letter(c)].width = w

    for r_idx, (_, row) in enumerate(df_best[[h for h, _ in best_headers]].iterrows(), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if c_idx == 6:
                cell.fill = conf_colors.get(val, PatternFill())

    ws1.auto_filter.ref = f"A1:{get_column_letter(len(best_headers))}{len(df_best)+1}"

    # --- Top 3 Candidates ---
    if not df_top3_out.empty:
        ws2 = wb.create_sheet('Top 3 Candidates')
        t3_headers = [
            (f'{LABEL_A}_Description', 55),
            (f'{LABEL_B}_Description', 55),
            ('TF-IDF_Score', 12),
            ('Fuzzy_Score', 12),
            ('Combined_Score', 14),
            ('Match_Rank', 12),
        ]
        for c, (h, w) in enumerate(t3_headers, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font, cell.fill = hfont, hfill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            ws2.column_dimensions[get_column_letter(c)].width = w

        for r_idx, (_, row) in enumerate(df_top3_out.iterrows(), 2):
            for c_idx, val in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=val).alignment = Alignment(
                    wrap_text=True, vertical='top')

        ws2.auto_filter.ref = f"A1:{get_column_letter(len(t3_headers))}{len(df_top3_out)+1}"

    # --- Unmatched ---
    ws3 = wb.create_sheet(f'Unmatched {LABEL_A}')
    matched_a = set(df_best[f'{LABEL_A}_Description']) if len(df_best) > 0 else set()
    unmatched = [d for d in descs_a if d not in matched_a]
    ws3.cell(row=1, column=1, value=f'Unmatched {LABEL_A} Descriptions').font = Font(bold=True)
    for i, d in enumerate(unmatched, 2):
        ws3.cell(row=i, column=1, value=d)
    ws3.column_dimensions['A'].width = 70

    wb.save(output_path)
    return df_best


# ============================================================
# MAIN
# ============================================================

def load_file(filepath, column, file_format):
    """Load descriptions from excel or csv."""
    if file_format == 'excel':
        df = pd.read_excel(filepath, engine='openpyxl')
    else:
        try:
            df = pd.read_csv(filepath, sep=CSV_SEP, encoding=CSV_ENCODING,
                             on_bad_lines='skip', engine='python')
        except TypeError:
            df = pd.read_csv(filepath, sep=CSV_SEP, encoding=CSV_ENCODING,
                             error_bad_lines=False, engine='python')
    return df[column].dropna().unique().tolist()


def main():
    print("=" * 60)
    print("Generic Description Matching Engine")
    print("=" * 60)

    # Load
    print(f"\nLoading {LABEL_A} from {FILE_A}...")
    descs_a = load_file(FILE_A, COL_A, FILE_A_FORMAT)
    print(f"  {len(descs_a)} unique descriptions")

    print(f"Loading {LABEL_B} from {FILE_B}...")
    descs_b = load_file(FILE_B, COL_B, FILE_B_FORMAT)
    print(f"  {len(descs_b)} unique descriptions")

    # Clean
    print("\nCleaning descriptions...")
    cleaner = DescriptionCleaner(expand_abbreviations=True, abbreviations_file=ABBREVIATIONS_FILE)

    patterns_a = cleaner.detect_patterns(descs_a)
    print(f"  {LABEL_A} detected patterns: "
          + ", ".join(k for k, v in patterns_a.items() if v))

    patterns_b = cleaner.detect_patterns(descs_b)
    print(f"  {LABEL_B} detected patterns: "
          + ", ".join(k for k, v in patterns_b.items() if v))

    descs_a_clean = [cleaner.clean(d) for d in descs_a]
    descs_b_clean = [cleaner.clean(d) for d in descs_b]

    # Remove empty descriptions after cleaning
    valid_a = [(i, d, c) for i, (d, c) in enumerate(zip(descs_a, descs_a_clean)) if c]
    valid_b = [(i, d, c) for i, (d, c) in enumerate(zip(descs_b, descs_b_clean)) if c]

    descs_a_orig = [d for _, d, _ in valid_a]
    descs_a_clean = [c for _, _, c in valid_a]
    descs_b_orig = [d for _, d, _ in valid_b]
    descs_b_clean = [c for _, _, c in valid_b]

    print(f"  After cleaning: {LABEL_A}={len(descs_a_clean)}, {LABEL_B}={len(descs_b_clean)}")

    # Match
    print("\nMatching descriptions...")
    matcher = DescriptionMatcher()
    candidates = matcher.compute_matches(descs_a_clean, descs_b_clean)

    print("\nAssigning 1:1 unique matches...")
    matches = matcher.greedy_1to1(candidates)
    print(f"  Final matched pairs: {len(matches)}")

    # Write output
    print(f"\nWriting results to {OUTPUT_FILE}...")
    df_best = write_results_excel(
        descs_a_orig, descs_b_orig,
        descs_a_clean, descs_b_clean,
        matches, candidates, OUTPUT_FILE
    )

    # Print summary
    if len(df_best) > 0:
        print(f"\n{'=' * 50}")
        print("RESULTS")
        print(f"{'=' * 50}")
        print(df_best['Confidence'].value_counts().to_string())
        print(f"\nTotal matched: {len(df_best)}")
        print(f"Unmatched {LABEL_A}: {len(descs_a_orig) - len(df_best)}")

        print(f"\nTop 10 matches:")
        for _, r in df_best.head(10).iterrows():
            a_desc = r[f'{LABEL_A}_Description']
            b_desc = r[f'{LABEL_B}_Description']
            print(f"  [{r['Combined_Score']:.0f} {r['Confidence']}] "
                  f"{str(a_desc)[:40]} <-> {str(b_desc)[:40]}")
    else:
        print("\nNo matches found above threshold.")

    print("\nDone!")


if __name__ == '__main__':
    main()
