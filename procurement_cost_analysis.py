"""
Procurement Cost Analysis: UMass vs Benchmark
==============================================
Matches item descriptions between UMass and a benchmark institution,
compares unit costs, and generates a formatted Excel report showing
where UMass is overpaying, saving, or on par.

Requirements:
    pip install pandas openpyxl rapidfuzz scikit-learn numpy

Usage:
    1. Update the file paths and column names under "CONFIGURE" below
    2. Run:  python procurement_cost_analysis.py
"""

import pandas as pd
import re
import numpy as np
from rapidfuzz import fuzz
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import warnings
warnings.filterwarnings('ignore')


# ============================================================
# CONFIGURE THESE
# ============================================================
UMASS_FILE = "umass_vendor_description_spend_v4.txt"
BENCHMARK_FILE = "cchs_vendor_description_spend_v3.txt"
OUTPUT_FILE = "umass_procurement_analysis.xlsx"

# File format
FILE_SEP = "|"                          # Delimiter (pipe for these files)
FILE_ENCODING = "utf-8"

# Column names in UMass file
U_DESCRIPTION = "Line_Description"
U_UNIT_COST = "UNIT_COST"
U_QUANTITY = "QUANTITY"
U_EXTENDED_AMT = "EXTENDED_AMOUNT"
U_UOM = "UNIT_OF_MEASURE"
U_ID_COL = "seq_no"                     # Any unique column for counting txns

# Column names in Benchmark file
B_DESCRIPTION = "LINE_DESCRIPTION"
B_UNIT_COST = "UNIT_COST"
B_QUANTITY = "QUANTITY"
B_EXTENDED_AMT = "EXTENDED_AMOUNT"
B_UOM = "UNIT_OF_MEASURE"
B_ID_COL = "SEQ_NO"

# Rows to exclude (regex, case-insensitive)
EXCLUDE_PATTERNS = r'FREIGHT|S/H|RESTOCKING'

# Matching thresholds
TFIDF_MIN_SCORE = 0.15        # Min TF-IDF cosine similarity
COMBINED_MIN_SCORE = 35       # Min combined score to keep a match
TFIDF_WEIGHT = 0.5
FUZZY_WEIGHT = 0.5

# Cost comparison thresholds
OVERPAY_THRESHOLD = 10        # % above benchmark = overpaying
SAVING_THRESHOLD = -10        # % below benchmark = saving
# ============================================================


def clean_desc(text):
    """Normalize description text for matching."""
    text = str(text).upper()
    text = re.sub(r'\[\s*CATALOG\s*:.*?\]', '', text)
    text = re.sub(r'\[\s*ITEM\s*:.*?\]', '', text)
    text = re.sub(r'ITM-\d+\s*-\s*', '', text)
    text = re.sub(r'\d{5,}\|', '', text)
    text = re.sub(r'DIV\d+/VARIOUS COMPONENTS:\s*', '', text)
    text = re.sub(r'[^\w\s/\.\-]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def load_and_aggregate(filepath, desc_col, cost_col, qty_col, amt_col, uom_col, id_col):
    """Load transaction data, filter, and aggregate by description."""
    df = pd.read_csv(filepath, sep=FILE_SEP, encoding=FILE_ENCODING,
                     error_bad_lines=False, engine='python')

    print(f"  Raw rows: {len(df)}")

    # Exclude freight/non-product rows
    mask = ~df[desc_col].str.upper().str.contains(EXCLUDE_PATTERNS, na=False)
    df = df[mask].copy()

    # Keep only positive unit costs
    df = df[df[cost_col] > 0].copy()
    print(f"  After filtering: {len(df)} rows, {df[desc_col].nunique()} unique items")

    # Aggregate by description
    agg = df.groupby(desc_col).agg(
        Avg_Unit_Cost=(cost_col, 'mean'),
        Med_Unit_Cost=(cost_col, 'median'),
        Min_Unit_Cost=(cost_col, 'min'),
        Max_Unit_Cost=(cost_col, 'max'),
        Total_Qty=(qty_col, 'sum'),
        Total_Spend=(amt_col, 'sum'),
        Txn_Count=(id_col, 'count'),
        UOM=(uom_col, 'first')
    ).reset_index()

    return agg


def match_descriptions(u_descs, b_descs):
    """Match UMass descriptions to benchmark using TF-IDF + fuzzy matching.
    Returns list of (u_idx, b_idx, tfidf_score, fuzzy_score, combined_score)."""

    u_clean = [clean_desc(d) for d in u_descs]
    b_clean = [clean_desc(d) for d in b_descs]

    # Build TF-IDF character n-gram model
    all_descs = u_clean + b_clean
    vectorizer = TfidfVectorizer(
        analyzer='char_wb', ngram_range=(3, 5),
        max_features=50000, sublinear_tf=True
    )
    tfidf_matrix = vectorizer.fit_transform(all_descs)
    u_matrix = tfidf_matrix[:len(u_clean)]
    b_matrix = tfidf_matrix[len(u_clean):]

    # Compute cosine similarity
    sim = cosine_similarity(u_matrix, b_matrix)

    # Greedy 1:1 matching (highest score first, no repeats)
    candidates = []
    for i in range(len(u_clean)):
        best_j = np.argmax(sim[i])
        candidates.append((sim[i][best_j], i, best_j))

    candidates.sort(reverse=True)

    used_u, used_b = set(), set()
    matches = []

    for score, i, j in candidates:
        if i not in used_u and j not in used_b:
            fz = fuzz.token_sort_ratio(u_clean[i], b_clean[j])
            combined = TFIDF_WEIGHT * score * 100 + FUZZY_WEIGHT * fz
            if combined >= COMBINED_MIN_SCORE:
                matches.append((i, j, round(score * 100, 1), round(fz, 1), round(combined, 1)))
                used_u.add(i)
                used_b.add(j)

    return matches


def build_comparison(u_agg, b_agg, matches, u_desc_col, b_desc_col):
    """Build comparison DataFrame from matched pairs."""
    rows = []
    for i, j, tfidf_s, fz_s, comb_s in matches:
        u_row = u_agg.iloc[i]
        b_row = b_agg.iloc[j]

        u_cost = u_row['Avg_Unit_Cost']
        b_cost = b_row['Avg_Unit_Cost']
        diff = u_cost - b_cost
        pct = (diff / b_cost * 100) if b_cost > 0 else 0

        if pct > OVERPAY_THRESHOLD:
            status = 'Overpaying'
        elif pct < SAVING_THRESHOLD:
            status = 'Saving'
        else:
            status = 'On Par'

        rows.append({
            'UMass_Description': u_row[u_desc_col],
            'Benchmark_Description': b_row[b_desc_col],
            'Match_Score': comb_s,
            'UMass_Avg_Unit_Cost': round(u_cost, 2),
            'Benchmark_Avg_Unit_Cost': round(b_cost, 2),
            'Cost_Difference': round(diff, 2),
            'Cost_Diff_Pct': round(pct, 1),
            'Status': status,
            'UMass_UOM': u_row['UOM'],
            'Benchmark_UOM': b_row['UOM'],
            'UMass_Total_Qty': u_row['Total_Qty'],
            'Benchmark_Total_Qty': b_row['Total_Qty'],
            'UMass_Total_Spend': round(u_row['Total_Spend'], 2),
            'Benchmark_Total_Spend': round(b_row['Total_Spend'], 2),
            'UMass_Txn_Count': u_row['Txn_Count'],
            'Benchmark_Txn_Count': b_row['Txn_Count'],
            'Potential_Savings': round(max(0, diff * u_row['Total_Qty']), 2),
        })

    df = pd.DataFrame(rows)

    # Flag UOM matches
    df['UOM_Match'] = df.apply(
        lambda r: str(r['UMass_UOM']).upper().strip() == str(r['Benchmark_UOM']).upper().strip(),
        axis=1
    )

    return df


def write_excel(df, output_path):
    """Write formatted Excel report with two sheets."""
    wb = Workbook()

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side('thin'), right=Side('thin'),
        top=Side('thin'), bottom=Side('thin')
    )
    status_colors = {
        'Overpaying': PatternFill('solid', fgColor='FFC7CE'),
        'Saving': PatternFill('solid', fgColor='C6EFCE'),
        'On Par': PatternFill('solid', fgColor='D9D9D9'),
    }

    # --- Sheet 1: Fair comparison (same UOM) ---
    ws1 = wb.active
    ws1.title = 'Cost Comparison (Same UOM)'

    df_fair = df[df['UOM_Match']].sort_values('Cost_Diff_Pct', ascending=False)

    columns = [
        ('UMass_Description',       'UMass Description',    50),
        ('Benchmark_Description',   'Benchmark Description', 50),
        ('UMass_Avg_Unit_Cost',     'UMass Avg Cost',       14),
        ('Benchmark_Avg_Unit_Cost', 'Benchmark Avg Cost',   14),
        ('Cost_Difference',         'Cost Diff ($)',        14),
        ('Cost_Diff_Pct',           'Cost Diff (%)',        14),
        ('Status',                  'Status',               14),
        ('Potential_Savings',       'Potential Savings ($)', 16),
        ('Match_Score',             'Match Score',          12),
        ('UMass_Total_Qty',         'UMass Qty',            12),
        ('UMass_UOM',               'UOM',                  10),
    ]

    for c, (_, display, width) in enumerate(columns, 1):
        cell = ws1.cell(row=1, column=c, value=display)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws1.column_dimensions[chr(64 + c) if c <= 26 else ''].width = width

    # Set column widths using openpyxl utility
    from openpyxl.utils import get_column_letter
    for c, (_, _, width) in enumerate(columns, 1):
        ws1.column_dimensions[get_column_letter(c)].width = width

    for r, (_, row) in enumerate(df_fair.iterrows(), 2):
        for c, (col, _, _) in enumerate(columns, 1):
            cell = ws1.cell(row=r, column=c, value=row[col])
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col == 'Status':
                cell.fill = status_colors.get(row[col], PatternFill())

    ws1.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(df_fair) + 1}"

    # --- Sheet 2: UOM mismatch (needs manual review) ---
    ws2 = wb.create_sheet('UOM Mismatch (Review)')

    df_mis = df[~df['UOM_Match']].sort_values('Match_Score', ascending=False)

    mis_columns = [
        ('UMass_Description',       'UMass Description',    50),
        ('Benchmark_Description',   'Benchmark Description', 50),
        ('UMass_Avg_Unit_Cost',     'UMass Avg Cost',       15),
        ('Benchmark_Avg_Unit_Cost', 'Benchmark Avg Cost',   15),
        ('UMass_UOM',               'UMass UOM',            12),
        ('Benchmark_UOM',           'Benchmark UOM',        12),
        ('Match_Score',             'Match Score',          12),
    ]

    for c, (_, display, width) in enumerate(mis_columns, 1):
        cell = ws2.cell(row=1, column=c, value=display)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws2.column_dimensions[get_column_letter(c)].width = width

    for r, (_, row) in enumerate(df_mis.iterrows(), 2):
        for c, (col, _, _) in enumerate(mis_columns, 1):
            ws2.cell(row=r, column=c, value=row[col]).alignment = Alignment(
                wrap_text=True, vertical='top'
            )

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(mis_columns))}{len(df_mis) + 1}"

    # --- Sheet 3: Summary ---
    ws3 = wb.create_sheet('Summary')
    ws3.sheet_properties.tabColor = '1F4E79'

    ws3['A1'] = 'UMass Procurement Cost Analysis'
    ws3['A1'].font = Font(bold=True, size=14)

    df_fair_only = df[df['UOM_Match']]
    summary_data = [
        ('', ''),
        ('Total matched item pairs', len(df)),
        ('Same UOM (fair comparison)', len(df_fair_only)),
        ('UOM mismatch (needs review)', len(df) - len(df_fair_only)),
        ('', ''),
        ('SAME-UOM BREAKDOWN', ''),
        ('Overpaying items (>{} %)'.format(OVERPAY_THRESHOLD),
         int((df_fair_only['Status'] == 'Overpaying').sum())),
        ('Saving items (<{} %)'.format(SAVING_THRESHOLD),
         int((df_fair_only['Status'] == 'Saving').sum())),
        ('On par items (within ±{} %)'.format(OVERPAY_THRESHOLD),
         int((df_fair_only['Status'] == 'On Par').sum())),
        ('', ''),
        ('Total potential savings ($)',
         round(df_fair_only[df_fair_only['Status'] == 'Overpaying']['Potential_Savings'].sum(), 2)),
        ('Total current savings ($)',
         round(abs(df_fair_only[df_fair_only['Status'] == 'Saving']['Cost_Difference']
               .multiply(df_fair_only[df_fair_only['Status'] == 'Saving']['UMass_Total_Qty']).sum()), 2)),
    ]

    for r, (label, value) in enumerate(summary_data, 3):
        ws3[f'A{r}'] = label
        ws3[f'B{r}'] = value
        if label and label.isupper():
            ws3[f'A{r}'].font = Font(bold=True)

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 18

    wb.save(output_path)


def main():
    # --- Load and aggregate ---
    print("Loading UMass data...")
    u_agg = load_and_aggregate(
        UMASS_FILE, U_DESCRIPTION, U_UNIT_COST, U_QUANTITY,
        U_EXTENDED_AMT, U_UOM, U_ID_COL
    )

    print("Loading benchmark data...")
    b_agg = load_and_aggregate(
        BENCHMARK_FILE, B_DESCRIPTION, B_UNIT_COST, B_QUANTITY,
        B_EXTENDED_AMT, B_UOM, B_ID_COL
    )

    # --- Match descriptions ---
    print("Matching descriptions (TF-IDF + fuzzy)...")
    u_descs = u_agg[U_DESCRIPTION].tolist()
    b_descs = b_agg[B_DESCRIPTION].tolist()
    matches = match_descriptions(u_descs, b_descs)
    print(f"  Matched pairs: {len(matches)}")

    # --- Build comparison ---
    print("Building cost comparison...")
    df_comp = build_comparison(u_agg, b_agg, matches, U_DESCRIPTION, B_DESCRIPTION)

    # --- Print summary ---
    df_fair = df_comp[df_comp['UOM_Match']]
    print(f"\n{'='*50}")
    print(f"RESULTS (same UOM only - fair comparison)")
    print(f"{'='*50}")
    print(f"Total matched pairs:    {len(df_comp)}")
    print(f"Same UOM pairs:         {len(df_fair)}")
    print(f"UOM mismatch (review):  {len(df_comp) - len(df_fair)}")
    print(f"")
    print(f"Overpaying items:       {(df_fair['Status'] == 'Overpaying').sum()}")
    print(f"Saving items:           {(df_fair['Status'] == 'Saving').sum()}")
    print(f"On par items:           {(df_fair['Status'] == 'On Par').sum()}")
    print(f"")

    sav = df_fair[df_fair['Status'] == 'Overpaying']['Potential_Savings'].sum()
    curr = abs(df_fair[df_fair['Status'] == 'Saving']['Cost_Difference']
               .multiply(df_fair[df_fair['Status'] == 'Saving']['UMass_Total_Qty']).sum())
    print(f"Potential savings:      ${sav:,.2f}")
    print(f"Current savings:        ${curr:,.2f}")

    # --- Top overpaying ---
    print(f"\nTop 10 overpaying items:")
    top_over = df_fair[df_fair['Status'] == 'Overpaying'].nlargest(10, 'Potential_Savings')
    for _, r in top_over.iterrows():
        desc = r['UMass_Description'][:45]
        print(f"  {desc:<45}  UMass: ${r['UMass_Avg_Unit_Cost']:>8.2f}  "
              f"Bench: ${r['Benchmark_Avg_Unit_Cost']:>8.2f}  "
              f"Diff: {r['Cost_Diff_Pct']:>+7.1f}%  "
              f"Savings: ${r['Potential_Savings']:>9.2f}")

    # --- Top saving ---
    print(f"\nTop 10 saving items:")
    top_save = df_fair[df_fair['Status'] == 'Saving'].nsmallest(10, 'Cost_Diff_Pct')
    for _, r in top_save.iterrows():
        desc = r['UMass_Description'][:45]
        print(f"  {desc:<45}  UMass: ${r['UMass_Avg_Unit_Cost']:>8.2f}  "
              f"Bench: ${r['Benchmark_Avg_Unit_Cost']:>8.2f}  "
              f"Diff: {r['Cost_Diff_Pct']:>+7.1f}%")

    # --- Write Excel ---
    print(f"\nWriting Excel report to {OUTPUT_FILE}...")
    write_excel(df_comp, OUTPUT_FILE)
    print("Done!")


if __name__ == '__main__':
    main()
