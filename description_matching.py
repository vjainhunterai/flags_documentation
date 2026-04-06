"""
Description Matching: CCHS vs UMass Invoice Descriptions
=========================================================
Matches item descriptions across two customers using:
  1. TF-IDF character n-gram similarity (catches partial/abbreviated matches)
  2. Fuzzy token-sort ratio (handles word reordering)
  3. Combined score with confidence tiers

Requirements:
  pip install pandas openpyxl rapidfuzz scikit-learn numpy

Usage:
  1. Update the 3 file paths below under "CONFIGURE THESE PATHS"
  2. Run:  python description_matching.py
"""

import pandas as pd
import re
import numpy as np
from rapidfuzz import fuzz
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import warnings
warnings.filterwarnings('ignore')

# ============================================================
# CONFIGURE THESE PATHS
# ============================================================
CCHS_FILE = "cchs_desc.xlsx"                          # Path to CCHS file
UMASS_FILE = "Item_description_umass.xlsx"              # Path to UMass file
OUTPUT_FILE = "description_matching_results.xlsx"       # Output file path

CCHS_COL = "Line_desc_cchs"                            # Column name in CCHS file
UMASS_COL = "Item_description_umass"                    # Column name in UMass file

# Matching parameters (tune as needed)
TFIDF_THRESHOLD = 0.15      # Minimum TF-IDF cosine similarity to consider
TOP_K = 3                   # Number of top candidates per item
BATCH_SIZE = 500            # Batch size for similarity computation
TFIDF_WEIGHT = 0.5          # Weight for TF-IDF score in combined score
FUZZY_WEIGHT = 0.5          # Weight for fuzzy score in combined score

# Confidence thresholds
HIGH_THRESHOLD = 75
MEDIUM_THRESHOLD = 50
LOW_THRESHOLD = 35
# ============================================================


def clean_desc(text):
    """Normalize and clean description text for matching."""
    text = str(text).upper()
    # Remove UMass-specific prefixes/catalog references
    text = re.sub(r'\[\s*CATALOG\s*:.*?\]', '', text)
    text = re.sub(r'\[\s*ITEM\s*:.*?\]', '', text)
    text = re.sub(r'ITM-\d+\s*-\s*', '', text)
    text = re.sub(r'\d{5,}\|', '', text)
    text = re.sub(r'DIV\d+/VARIOUS COMPONENTS:\s*', '', text)
    # Remove special characters, keep alphanumeric + basic punctuation
    text = re.sub(r'[^\w\s/\.\-]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def confidence_tier(score):
    """Assign confidence label based on combined score."""
    if score >= HIGH_THRESHOLD:
        return 'High'
    if score >= MEDIUM_THRESHOLD:
        return 'Medium'
    if score >= LOW_THRESHOLD:
        return 'Low'
    return 'Very Low'


def main():
    # --- Load data ---
    print("Loading data...")
    df_cchs = pd.read_excel(CCHS_FILE)
    df_umass = pd.read_excel(UMASS_FILE)

    cchs_descs = df_cchs[CCHS_COL].dropna().unique().tolist()
    umass_descs = df_umass[UMASS_COL].dropna().unique().tolist()
    print(f"  CCHS descriptions:  {len(cchs_descs)}")
    print(f"  UMass descriptions: {len(umass_descs)}")

    # --- Clean descriptions ---
    print("Cleaning descriptions...")
    cchs_clean = [clean_desc(d) for d in cchs_descs]
    umass_clean = [clean_desc(d) for d in umass_descs]

    # --- Build TF-IDF model ---
    print("Building TF-IDF model (character n-grams)...")
    all_descs = cchs_clean + umass_clean
    vectorizer = TfidfVectorizer(
        analyzer='char_wb',
        ngram_range=(3, 5),
        max_features=50000,
        sublinear_tf=True
    )
    tfidf_matrix = vectorizer.fit_transform(all_descs)
    cchs_matrix = tfidf_matrix[:len(cchs_clean)]
    umass_matrix = tfidf_matrix[len(cchs_clean):]

    # --- Compute matches in batches ---
    print("Computing similarity matches...")
    results = []
    total = len(cchs_clean)

    for i in range(0, total, BATCH_SIZE):
        batch_end = min(i + BATCH_SIZE, total)
        sim = cosine_similarity(cchs_matrix[i:batch_end], umass_matrix)

        for j in range(sim.shape[0]):
            top_indices = np.argsort(sim[j])[-TOP_K:][::-1]
            top_scores = sim[j][top_indices]
            cchs_idx = i + j

            for rank, (umass_idx, score) in enumerate(zip(top_indices, top_scores)):
                if score > TFIDF_THRESHOLD:
                    fz_score = fuzz.token_sort_ratio(
                        cchs_clean[cchs_idx], umass_clean[umass_idx]
                    )
                    combined = TFIDF_WEIGHT * score * 100 + FUZZY_WEIGHT * fz_score
                    results.append({
                        'CCHS_Description': cchs_descs[cchs_idx],
                        'CCHS_Cleaned': cchs_clean[cchs_idx],
                        'UMass_Description': umass_descs[umass_idx],
                        'UMass_Cleaned': umass_clean[umass_idx],
                        'TF-IDF_Score': round(score * 100, 1),
                        'Fuzzy_Score': round(fz_score, 1),
                        'Combined_Score': round(combined, 1),
                        'Match_Rank': rank + 1
                    })

        processed = min(i + BATCH_SIZE, total)
        if processed % 2000 == 0 or processed == total:
            print(f"  Processed {processed}/{total}...")

    print(f"  Total candidate pairs: {len(results)}")

    # --- Aggregate results ---
    df_results = pd.DataFrame(results)

    # Best match per CCHS item
    df_best = df_results.loc[
        df_results.groupby('CCHS_Description')['Combined_Score'].idxmax()
    ].copy()
    df_best.sort_values('Combined_Score', ascending=False, inplace=True)
    df_best['Confidence'] = df_best['Combined_Score'].apply(confidence_tier)

    # Top-3 candidates per CCHS item
    df_top3 = (
        df_results
        .sort_values(['CCHS_Description', 'Combined_Score'], ascending=[True, False])
        .groupby('CCHS_Description').head(TOP_K)
        .reset_index(drop=True)
    )

    print(f"\nMatch Summary:")
    print(df_best['Confidence'].value_counts().to_string())
    print(f"Total CCHS items matched: {len(df_best)}")

    # --- Write Excel output ---
    print(f"\nWriting output to {OUTPUT_FILE}...")
    wb = Workbook()

    # -- Sheet 1: Summary --
    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = 'Description Matching Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'] = 'CCHS Descriptions'
    ws['B3'] = len(cchs_descs)
    ws['A4'] = 'UMass Descriptions'
    ws['B4'] = len(umass_descs)
    ws['A5'] = 'Matched (any confidence)'
    ws['B5'] = len(df_best)
    ws['A7'] = 'Confidence Breakdown'
    ws['A7'].font = Font(bold=True)
    for idx, conf in enumerate(['High', 'Medium', 'Low', 'Very Low'], 8):
        ws[f'A{idx}'] = conf
        ws[f'B{idx}'] = int((df_best['Confidence'] == conf).sum())

    # -- Sheet 2: Best Matches --
    ws_best = wb.create_sheet('Best Matches')
    best_headers = [
        'CCHS_Description', 'UMass_Description',
        'TF-IDF_Score', 'Fuzzy_Score', 'Combined_Score', 'Confidence'
    ]
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    conf_colors = {
        'High': PatternFill('solid', fgColor='C6EFCE'),
        'Medium': PatternFill('solid', fgColor='FFEB9C'),
        'Low': PatternFill('solid', fgColor='FFC7CE'),
        'Very Low': PatternFill('solid', fgColor='D9D9D9'),
    }

    for col, h in enumerate(best_headers, 1):
        cell = ws_best.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    for r_idx, (_, row_data) in enumerate(df_best[best_headers].iterrows(), 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_best.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if c_idx == 6:
                cell.fill = conf_colors.get(val, PatternFill())

    ws_best.column_dimensions['A'].width = 55
    ws_best.column_dimensions['B'].width = 55
    ws_best.column_dimensions['C'].width = 12
    ws_best.column_dimensions['D'].width = 12
    ws_best.column_dimensions['E'].width = 14
    ws_best.column_dimensions['F'].width = 12
    ws_best.auto_filter.ref = f"A1:F{len(df_best)+1}"

    # -- Sheet 3: Top 3 Candidates --
    ws_top3 = wb.create_sheet('Top 3 Candidates')
    t3_headers = [
        'CCHS_Description', 'UMass_Description',
        'TF-IDF_Score', 'Fuzzy_Score', 'Combined_Score', 'Match_Rank'
    ]
    for col, h in enumerate(t3_headers, 1):
        cell = ws_top3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for r_idx, (_, row_data) in enumerate(df_top3[t3_headers].iterrows(), 2):
        for c_idx, val in enumerate(row_data, 1):
            ws_top3.cell(row=r_idx, column=c_idx, value=val).alignment = Alignment(
                wrap_text=True, vertical='top'
            )

    ws_top3.column_dimensions['A'].width = 55
    ws_top3.column_dimensions['B'].width = 55
    ws_top3.column_dimensions['C'].width = 12
    ws_top3.column_dimensions['D'].width = 12
    ws_top3.column_dimensions['E'].width = 14
    ws_top3.column_dimensions['F'].width = 12
    ws_top3.auto_filter.ref = f"A1:F{len(df_top3)+1}"

    # -- Sheet 4: Unmatched CCHS --
    ws_unmatched = wb.create_sheet('Unmatched CCHS')
    unmatched = set(cchs_descs) - set(df_best['CCHS_Description'])
    ws_unmatched.cell(row=1, column=1, value='Unmatched CCHS Descriptions').font = Font(bold=True)
    for i, desc in enumerate(sorted(unmatched, key=str), 2):
        ws_unmatched.cell(row=i, column=1, value=desc)
    ws_unmatched.column_dimensions['A'].width = 70

    wb.save(OUTPUT_FILE)
    print(f"Done! Output saved to: {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
